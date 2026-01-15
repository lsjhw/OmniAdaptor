"""
   Copyright (c) Huawei Technologies Co., Ltd. 2026. All rights reserved.
   You can use this software according to the terms and conditions of the Mulan PSL v2.
   You may obtain a copy of Mulan PSL v2 at:
            http://license.coscl.org.cn/MulanPSL2
   THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
   EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
   MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
   See the Mulan PSL v2 for more details.
"""
import os
import time
import pandas as pd
import argparse
import subprocess
from tqdm import tqdm
from pathlib import Path
from datetime import datetime
from omnihelper.parser.function_parser import FunctionParser
from omnihelper.util.common_util import CommonUtil


class LogParser:

    EXECUTE_PATH = CommonUtil.get_execute_path()
    TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
    TMP_PATH = os.path.join(EXECUTE_PATH, "tmp", TIMESTAMP)

    def __init__(self):
        self.parser = None
        self.args = None
        self.compressed_files = []
        self.input_is_file = False  # 标记输入是否为文件
        self.input_file_path = None  # 如果是文件，存储文件路径

        self._create_parser()
        self._get_arguments()
        self.print_arguments()

    def _create_parser(self):
        self.parser = argparse.ArgumentParser(
            description='Big Data Operator Scanning Command Line Tool',
            formatter_class=argparse.RawDescriptionHelpFormatter,
            epilog="""
Usage Examples:
  ./omnihelper -i ./input_data -o ./output_dir 
    --java-path /path/to/java/bin/java 
    --class-path /path/to/boostkit-omnimv-logparser-spark-3.4.3-1.2.0-aarch64.jar:\
/path/to/boostkit-omnimv-spark-3.4.3-1.2.0-aarch64.jar:\
/path/to/spark-3.4.3-bin-hadoop3/jars/*

  ./omnihelper -i ./input_data/eventlog.lz4 -o ./output_dir
    --java-path /path/to/java/bin/java 
    --class-path /path/to/boostkit-omnimv-logparser-spark-3.4.3-1.2.0-aarch64.jar:\
/path/to/boostkit-omnimv-spark-3.4.3-1.2.0-aarch64.jar:\
/path/to/spark-3.4.3-bin-hadoop3/jars/*
"""
        )

        # 必需参数
        self.parser.add_argument(
            '--input_data', '-i',
            type=str,
            required=True,
            help='Input directory path or single file path (required). '
                 'If a single .lz4 or .zstd file is provided, only that file will be processed.'
        )

        # 可选参数
        self.parser.add_argument(
            '--output_dir', '-o',
            type=str,
            default=None,
            help='Output directory path (default: ./output)'
        )

        # Java相关参数组
        java_group = self.parser.add_argument_group('Java Configuration')

        # Java可执行文件路径
        java_group.add_argument(
            '--java-path',
            type=str,
            default="java",
            help='Java executable path (default: "java" from system PATH)'
        )

        # Java Class 路劲
        java_group.add_argument(
            '--class-path',
            type=str,
            required=True,
            help='Complete Java classpath string'
        )

    def _get_arguments(self):
        self.args = self.parser.parse_args()

        # 检查输入是文件还是目录
        if os.path.isfile(self.args.input_data):
            self.input_is_file = True
            self.input_file_path = os.path.realpath(self.args.input_data)

            # 检查文件扩展名
            filename_lower = os.path.basename(self.args.input_data).lower()
            if not (filename_lower.endswith('.lz4') or filename_lower.endswith('.zstd')):
                self.parser.error(
                    f"Input file must be a .lz4 or .zstd file. "
                    f"Provided file: {os.path.basename(self.args.input_data)}"
                )
        elif os.path.isdir(self.args.input_data):
            self.input_is_file = False
        else:
            self.parser.error(f"Input path is neither a file nor a directory: {self.args.input_data}")

        # 输出目录默认值
        if self.args.output_dir is None:
            self.args.output_dir = os.path.join(self.EXECUTE_PATH, "output")

        os.makedirs(self.args.output_dir, exist_ok=True)
        os.makedirs(self.TMP_PATH, exist_ok=True)

    def print_arguments(self):
        # 打印配置信息
        print("=" * 60)
        print("  Big Data Operator Scanning Tool")
        print("=" * 60)

        if self.input_is_file:
            print(f"Input File: {self.input_file_path}")
        else:
            print(f"Input Directory: {os.path.realpath(self.args.input_data)}")

        print(f"Output Directory: {os.path.realpath(self.args.output_dir)}")
        print(f"Temporary Directory: {os.path.realpath(self.TMP_PATH)}")
        print(f"Java Path: {self.args.java_path}")
        print(f"Java Class Path: {self.args.class_path}")
        print("-" * 60)

    def parse_single_file(self, input_file_path: str, output_file_path: str, filename: str):
        cmd = [
            self.args.java_path,
            "-cp",
            self.args.class_path,
            "org.apache.spark.deploy.history.ParseLog",
            input_file_path,
            output_file_path,
            filename
        ]
        try:
            process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, encoding='gbk')
            stdout, stderr = process.communicate()
            return process.returncode == 0, stdout, stderr
        except Exception as e:
            return False, "", str(e)

    def find_compressed_files(self):
        """查找目录/文件中的所有.lz4和.zstd文件"""
        if self.input_is_file:
            # 处理单个文件
            input_file_path = os.path.dirname(self.input_file_path)
            filename = os.path.basename(self.input_file_path)
            output_file_path = self.TMP_PATH

            self.compressed_files.append({
                "input_file_path": input_file_path,
                "output_file_path": output_file_path,
                "filename": filename
            })
        else:
            # 处理目录
            for root, _, files in os.walk(self.args.input_data):
                for filename in files:
                    if filename.lower().endswith(('.lz4', '.zstd')):
                        input_file_path = Path(root)
                        output_file_path = self.TMP_PATH / Path(root).relative_to(self.args.input_data)
                        self.compressed_files.append({
                            "input_file_path": os.path.realpath(input_file_path),
                            "output_file_path": os.path.realpath(output_file_path),
                            "filename": filename
                        })

    def get_execution_plan(self):
        failed_files = []  # 存储失败的文件信息
        print("Start parsing event log...")

        with tqdm(total=len(self.compressed_files), desc="Processing ") as pbar:
            for compressed_file in self.compressed_files:
                input_file_path = compressed_file["input_file_path"]
                output_file_path = compressed_file["output_file_path"]
                filename = compressed_file["filename"]
                pbar.set_description(f"Processing: {filename[:40]}{'...' if len(filename) > 40 else ''}")
                os.makedirs(output_file_path, exist_ok=True)
                success, stdout, stderr = self.parse_single_file(input_file_path, output_file_path, filename)
                if not success:
                    failed_files.append((os.path.join(input_file_path, filename), stderr))
                pbar.update(1)

        # 统一展示失败结果
        if failed_files:
            print(f"Processing completed with {len(failed_files)} failed files:")
            for i, (filepath, error) in enumerate(failed_files, 1):
                print(f"{i}. [File]: {filepath}")
                print(f"   [Error]: {error}")
            print(f"\nSummary: {len(self.compressed_files) - len(failed_files)} succeeded, {len(failed_files)} failed")
        else:
            print(f"\nAll {len(self.compressed_files)} files processed successfully!")
        print("-" * 60)

    def parse_event_log(self):
        try:
            records = []
            expr_parser = FunctionParser(self.TMP_PATH)
            start_time = time.time()
            res = expr_parser.parse_event_log()
            print("Cost time: " + str(time.time() - start_time))
            for item in res:
                for app_id, app_info_list in item.items():
                    for app_info in app_info_list:
                        # 提取基本信息
                        func_name = app_info.get('func_name', '')
                        sql_hash = app_info.get('sql_hash', '')
                        inputs = app_info.get('input', [])
                        times = app_info.get('times', 0)
                        record = {
                            'ApplicationID+SQL ID': app_id,
                            'SQL Hash': sql_hash,
                            'Omni不支持的算子名称': '',
                            'Omni不支持的算子Input': '',
                            'Omni不支持的算子Output': '',
                            'Omni不支持的算子出现频次': '',
                            'Omni不支持的算子Output rows': '',
                            'Omni不支持的表达式/内置函数名称': func_name,
                            'Omni不支持的表达式/内置函数Input': ",".join(inputs),
                            'Omni不支持的表达式/内置函数出现频次': times,
                            'Spark版本': '',  # 原始数据中没有Spark版本信息
                            '异常信息/备注': ''  # 添加函数名作为备注
                        }
                        records.append(record)
            df = pd.DataFrame(records)
            output_excel_path = os.path.join(self.args.output_dir, f"Omni_Analysis_Report_{self.TIMESTAMP}.xlsx")
            # 使用ExcelWriter来添加格式
            with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
                # 先写入数据
                df.to_excel(writer, index=False, header=False ,startrow=2)

                # 获取工作簿和工作表
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                # 添加两行标题
                # 第一行标题
                worksheet.cell(row=1, column=1, value='ApplicationID+SQL ID')
                worksheet.cell(row=1, column=2, value='SQL Hash')
                worksheet.cell(row=1, column=3, value='Omni不支持的算子')
                worksheet.cell(row=1, column=8, value='Omni不支持的表达式/内置函数')
                worksheet.cell(row=1, column=11, value='Spark版本')
                worksheet.cell(row=1, column=12, value='异常信息/备注')

                # 第二行标题（子标题）
                column_titles = [
                    'ApplicationID+SQL ID', 'SQL Hash',
                    '名称', 'Input', 'Output', '出现频次', 'Output rows',
                    '名称', 'Input', '出现频次',
                    'Spark版本', '异常信息/备注'
                ]

                for col_idx, title in enumerate(column_titles, 1):
                    worksheet.cell(row=2, column=col_idx, value=title)

                # 合并单元格
                # 第1列：第1-2行合并
                worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
                # 第2列：第1-2行合并
                worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
                # 第3-7列：第1行合并（Omni不支持的算子）
                worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=7)
                # 第8-10列：第1行合并（Omni不支持的表达式/内置函数）
                worksheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=10)
                # 第11列：第1-2行合并
                worksheet.merge_cells(start_row=1, start_column=11, end_row=2, end_column=11)
                # 第12列：第1-2行合并
                worksheet.merge_cells(start_row=1, start_column=12, end_row=2, end_column=12)

                # 导入样式
                from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

                # 定义样式
                center_alignment = Alignment(horizontal='center', vertical='center')
                bold_font = Font(bold=True)
                header_fill = PatternFill(start_color="EBEFF6", end_color="EBEFF6", fill_type="solid")

                # 定义边框样式
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # 计算总行数和列数
                total_rows = len(df) + 2  # 数据行数 + 标题行
                total_cols = len(column_titles)

                # 先为所有单元格应用样式（包括合并单元格）
                for row in range(1, total_rows + 1):
                    for col in range(1, total_cols + 1):
                        cell = worksheet.cell(row=row, column=col)

                        # 所有单元格应用边框
                        cell.border = thin_border

                        # 前两行（标题行）应用特殊样式
                        if row <= 2:
                            cell.font = bold_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                        else:
                            # 数据行应用居中对齐
                            cell.alignment = center_alignment

                # 使用固定的列宽设置
                from openpyxl.utils import get_column_letter

                column_widths = [35, 10, 20, 35, 35, 10, 10, 20, 35, 10, 15, 30]

                # 为每列设置固定的宽度
                for col_idx, width in enumerate(column_widths, 1):
                    column_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[column_letter].width = width

                # 对于第1,2列值相同的行，合并单元格
                # 数据从第3行开始
                start_data_row = 3
                current_start_row = start_data_row

                # 遍历所有数据行
                for current_row in range(start_data_row, total_rows + 1):
                    # 获取当前行的第1列和第2列值
                    current_col1_value = worksheet.cell(row=current_row, column=1).value
                    current_col2_value = worksheet.cell(row=current_row, column=2).value

                    # 获取下一行的第1列和第2列值（如果存在）
                    if current_row < total_rows:
                        next_col1_value = worksheet.cell(row=current_row + 1, column=1).value
                        next_col2_value = worksheet.cell(row=current_row + 1, column=2).value
                    else:
                        next_col1_value = None
                        next_col2_value = None

                    # 检查当前行是否与下一行的第1、2列值相同
                    same_values = (current_col1_value == next_col1_value and
                                   current_col2_value == next_col2_value)

                    # 如果值不相同或者已经到最后一行，则合并之前相同的行
                    if not same_values:
                        # 如果有多行相同的值，合并它们
                        if current_start_row < current_row:
                            # 合并第1列
                            worksheet.merge_cells(
                                start_row=current_start_row,
                                start_column=1,
                                end_row=current_row,
                                end_column=1
                            )
                            # 合并第2列
                            worksheet.merge_cells(
                                start_row=current_start_row,
                                start_column=2,
                                end_row=current_row,
                                end_column=2
                            )

                            # 设置合并后的单元格样式
                            for col in [1, 2]:
                                merged_cell = worksheet.cell(row=current_start_row, column=col)
                                merged_cell.alignment = center_alignment
                                merged_cell.border = thin_border

                        # 更新起始行为下一行
                        current_start_row = current_row + 1

            print(f"[SUCCESS] Analysis report has been saved to: {output_excel_path}")

        except Exception as e:
            print(f"[ERROR] ：{e}")
        print("-" * 60)


def main():
    logparser = LogParser()
    logparser.find_compressed_files()
    logparser.get_execution_plan()
    logparser.parse_event_log()


if __name__ == "__main__":
    main()