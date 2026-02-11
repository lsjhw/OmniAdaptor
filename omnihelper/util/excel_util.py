"""
   Copyright (c) Huawei Technologies Co., Ltd. 2026. All rights reserved.
   You can use this software according to the terms and conditions of the Mulan PSL v2.
   You may obtain a copy of Mulan PSL v2 at:
            http://license.coscl.org.cn/MulanPSL2
   THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
   EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
   MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
   See the Mulan PSL v2 for more details.
"""

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelWriterWithStyle:
    """带样式的Excel写入器"""

    # 常量定义
    DEFAULT_START_ROW = 3
    DEFAULT_VALUE = True

    def __init__(self):
        # 定义表头结构（使用字典结构更灵活）
        # main：主标题  sub:副标题  merge_start:该部分合并起始点  merged_op_column 同名算子合并项 show_detail 展示算子细节
        self.headers = [
            {'main': 'ApplicationID+SQL ID', "column_width": 35},
            {'main': 'SQL Hash', "column_width": 10},
            {'main': 'Omni不支持的算子', 'sub': '名称', 'merge_start': self.DEFAULT_VALUE,
             "column_width": 20, "merged_op_column": self.DEFAULT_VALUE},
            {'main': 'Omni不支持的算子', 'sub': 'Input', "column_width": 35},
            {'main': 'Omni不支持的算子', 'sub': 'Output', "column_width": 35},
            {'main': 'Omni不支持的算子', 'sub': '出现频次', "column_width": 10},
            {'main': 'Omni不支持的算子', 'sub': '运行时间（s）', "column_width": 35,
             "merged_op_column": self.DEFAULT_VALUE},
            {'main': 'Omni不支持的算子', 'sub': '文件大小（MiB）', "column_width": 15,
             "merged_op_column": self.DEFAULT_VALUE, "show_details": self.DEFAULT_VALUE},
            {'main': 'Omni不支持的算子', 'sub': 'Output rows', "column_width": 15,
             "merged_op_column": self.DEFAULT_VALUE, "show_details": self.DEFAULT_VALUE},
            {'main': 'Omni不支持的表达式/内置函数', 'sub': '名称', 'merge_start': self.DEFAULT_VALUE,
             "column_width": 20},
            {'main': 'Omni不支持的表达式/内置函数', 'sub': 'Input', "column_width": 35},
            {'main': 'Omni不支持的表达式/内置函数', 'sub': '出现频次', "column_width": 10},
            {'main': 'Omni不支持的表达式/内置函数', 'sub': '是否udf', "column_width": 10},
            {'main': 'Spark版本', "column_width": 15},
            {'main': '异常信息/备注', "column_width": 30}
        ]
        self.merged_op_column = []

        # 定义样式
        self._init_styles()

    def _init_styles(self):
        """初始化样式设置"""
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.bold_font = Font(bold=True)
        self.header_fill = PatternFill(start_color="EBEFF6", end_color="EBEFF6", fill_type="solid")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def write_to_excel(self, df, output_excel_path, show_op_details, sheet_name='Sheet1'):
        """
        将DataFrame写入Excel并应用样式

        Parameters:
        -----------
        df : pandas.DataFrame
            要写入的数据
        output_excel_path : str
            输出Excel文件路径
        sheet_name : str, optional
            工作表名称，默认为'Sheet1'
        """
        print("Start writing to excel...")

        try:

            if not show_op_details:
                self.headers = [i for i in self.headers if "show_details" not in i]

            # 确保DataFrame列数与配置一致
            if len(df.columns) != len(self.headers):
                print(f"[ERROR] DataFrame has {len(df.columns)} columns, "
                      f"but config has {len(self.headers)} columns. ")
                return False

            with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
                # 先写入数据
                df.to_excel(writer, index=False, header=False, startrow=2)

                # 获取工作簿和工作表
                worksheet = writer.sheets[sheet_name]

                # 应用样式
                self._apply_styles(worksheet, df)

                # 设置标题及列宽
                self._write_headers(worksheet)

                # 合并相同值的单元格
                self._merge_duplicate_cells(worksheet, df)

            print(f"[SUCCESS] Analysis report has been saved to: {output_excel_path}")
            return True

        except Exception as e:
            print(f"[ERROR] Failed to write Excel file: {e}")
            return False

    def _write_headers(self, worksheet):
        """
        写入表头到指定的工作表中。

        参数:
        worksheet: 工作表对象，用于写入表头信息。

        该方法根据self.headers中的信息，将主标题和子标题写入工作表的指定位置，
        并根据需要合并单元格。
        """
        # 遍历self.headers中的每个表头项，enumerate从1开始计数，对应Excel列号
        for col_idx, item in enumerate(self.headers, 1):
            # 如果表头项中指定了列宽，则设置对应列的宽度
            if 'column_width' in item:
                column_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[column_letter].width = item['column_width']

            # 如果表头项有主标题，并且没有子标题或没有指定合并结束位置，则写入主标题到第一行
            if 'main' in item and ('sub' not in item or 'merge_start' in item):
                worksheet.cell(row=1, column=col_idx, value=item['main'])
            # 如果表头项有子标题，则写入子标题到第二行
            if 'sub' in item:
                worksheet.cell(row=2, column=col_idx, value=item['sub'])

            # 如果表头项只有主标题没有子标题，则合并第一行和第二行的对应列
            if 'main' in item and 'sub' not in item:
                worksheet.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            # 如果表头项有主标题、子标题，并且指定了合并结束位置，则合并主标题所在的单元格
            if 'main' in item and 'sub' in item and 'merge_start' in item:
                merge_end = len([i for i in self.headers if i['main'] == item['main']]) + col_idx - 1
                worksheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=merge_end)

            # 如果表头项指定了需要合并操作的列，则将该列索引添加到merged_op_column列表中
            if 'merged_op_column' in item and item['merged_op_column']:
                self.merged_op_column.append(col_idx)

    def _apply_styles(self, worksheet, df):
        """应用样式（自动处理所有列）"""
        total_rows = len(df) + 2  # 数据行数 + 标题行
        total_cols = len(self.headers)

        # 为所有单元格应用样式
        for row in range(1, total_rows + 1):
            for col in range(1, total_cols + 1):
                cell = worksheet.cell(row=row, column=col)

                # 设置单元格的数据类型为字符串
                cell.data_type = 's'
                cell.number_format = '@'

                # 确保单元格值为字符串
                if cell.value is not None and not isinstance(cell.value, str):
                    cell.value = str(cell.value)

                # 所有单元格应用边框
                cell.border = self.thin_border

                # 前两行（标题行）应用特殊样式
                if row <= 2:
                    cell.font = self.bold_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_alignment
                else:
                    # 数据行应用居中对齐
                    cell.alignment = self.center_alignment

    def _merge_duplicate_cells(self, worksheet, df):
        """合并相同值的单元格"""
        title_start_row = self.DEFAULT_START_ROW
        op_start_row = self.DEFAULT_START_ROW
        total_rows = len(df) + 2

        # 遍历所有数据行
        for current_row in range(self.DEFAULT_START_ROW, total_rows + 1):
            current_col1_value = worksheet.cell(row=current_row, column=1).value
            current_col2_value = worksheet.cell(row=current_row, column=2).value
            current_col3_value = worksheet.cell(row=current_row, column=3).value

            # 获取下一行的值（如果存在）
            if current_row < total_rows:
                next_col1_value = worksheet.cell(row=current_row + 1, column=1).value
                next_col2_value = worksheet.cell(row=current_row + 1, column=2).value
                next_col3_value = worksheet.cell(row=current_row + 1, column=3).value
            else:
                next_col1_value = None
                next_col2_value = None
                next_col3_value = None

            # 检查当前行是否与下一行的第1、2列值相同
            same_values = (current_col1_value == next_col1_value and
                           current_col2_value == next_col2_value)

            if not same_values or not current_col3_value == next_col3_value:
                if op_start_row < current_row:
                    # 设置合并后的单元格样式
                    for col in self.merged_op_column:
                        worksheet.merge_cells(start_row=op_start_row, start_column=col, end_row=current_row,
                                              end_column=col)
                        merged_cell = worksheet.cell(row=op_start_row, column=col)
                        merged_cell.alignment = self.center_alignment
                        merged_cell.border = self.thin_border
                # 更新起始行为下一行
                op_start_row = current_row + 1

            # 如果值不相同或者已经到最后一行，则合并之前相同的行
            if not same_values:
                if title_start_row < current_row:
                    # 合并第1列
                    worksheet.merge_cells(start_row=title_start_row, start_column=1, end_row=current_row, end_column=1)
                    # 合并第2列
                    worksheet.merge_cells(start_row=title_start_row, start_column=2, end_row=current_row, end_column=2)

                    # 设置合并后的单元格样式
                    for col in [1, 2]:
                        worksheet.merge_cells(start_row=title_start_row, start_column=col, end_row=current_row,
                                              end_column=col)
                        merged_cell = worksheet.cell(row=title_start_row, column=col)
                        merged_cell.alignment = self.center_alignment
                        merged_cell.border = self.thin_border

                # 更新起始行为下一行
                title_start_row = current_row + 1