"""
   Copyright (c) Huawei Technologies Co., Ltd. 2026. All rights reserved.
   You can use this software according to the terms and conditions of the Mulan PSL v2.
   You may obtain a copy of Mulan PSL v2 at:
            http://license.coscl.org.cn/MulanPSL2
   THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
   EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
   MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
   See the Mulan PSL v2 for more details.
"""

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class FlinkExcelWriterWithStyle:
    """带样式的Flink Excel写入器"""

    # 常量定义
    DEFAULT_START_ROW = 3
    DEFAULT_VALUE = True

    def __init__(self):
        # 定义表头结构（使用字典结构更灵活）
        # main：主标题  sub:副标题  merge_start:该部分合并起始点  need_merged_column 需要合并项
        self.headers = [
            {'main': 'jobid', "column_width": 30, "need_merged_column": self.DEFAULT_VALUE},
            {'main': 'taskid', "column_width": 30, "need_merged_column": self.DEFAULT_VALUE},
            {'main': '状态', "column_width": 15},
            {'main': '算子名称', "column_width": 25},
            {'main': 'Input', "column_width": 30},
            {'main': 'Output', "column_width": 30},
            {'main': '出现频次', "column_width": 10},
            {'main': '运行时间(s)', "column_width": 15},
            {'main': '输入数据量', "column_width": 15},
            {'main': '输出数据量', "column_width": 15},
            {'main': '表达式/内置函数名称', "column_width": 30},
            {'main': 'Input', "column_width": 30},
            {'main': '嵌套内容', "column_width": 40},
            {'main': '出现频次', "column_width": 10}
        ]
        self.need_merged_column = []

        # 定义样式
        self._init_styles()

    def _init_styles(self):
        """初始化样式设置"""
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.bold_font = Font(bold=True)
        self.header_fill = PatternFill(start_color="EBEFF6", end_color="EBEFF6", fill_type="solid")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def write_to_excel(self, df, output_excel_path, sheet_name='Sheet1'):
        """
        将DataFrame写入Excel并应用样式

        Parameters:
        -----------
        df : pandas.DataFrame
            要写入的数据
        output_excel_path : str
            输出Excel文件路径
        sheet_name : str, optional
            工作表名称，默认为'Sheet1'
        """
        print("Start writing to excel...")

        try:
            # 确保DataFrame列数与配置一致
            if len(df.columns) != len(self.headers):
                print(f"[ERROR] DataFrame has {len(df.columns)} columns, "
                      f"but config has {len(self.headers)} columns. ")
                return False

            with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
                # 先写入数据
                df.to_excel(writer, index=False, header=False, startrow=2)

                # 获取工作簿和工作表
                worksheet = writer.sheets[sheet_name]

                # 应用样式
                self._apply_styles(worksheet, df)

                # 设置标题及列宽
                self._write_headers(worksheet)

                # 只合并 jobid 和 taskid 列，其他列都不合并
                self.merge_cells_full(
                    worksheet,
                    start_row=3,
                    independent_cols=[1],  # jobid和taskid需要合并
                    linked_cols=[2],  # 其他列都不合并
                    control_cols=[1]  # 不需要控制列
                )

            print(f"[SUCCESS] Analysis report has been saved to: {output_excel_path}")
            return True

        except Exception as e:
            print(f"[ERROR] Failed to write Excel file: {e}")
            return False

    def _write_headers(self, worksheet):
        """
        写入表头到指定的工作表中。

        参数:
        worksheet: 工作表对象，用于写入表头信息。

        该方法根据self.headers中的信息，将主标题和子标题写入工作表的指定位置，
        并根据需要合并单元格。
        """
        # 遍历self.headers中的每个表头项，enumerate从1开始计数，对应Excel列号
        for col_idx, item in enumerate(self.headers, 1):
            # 如果表头项中指定了列宽，则设置对应列的宽度
            if 'column_width' in item:
                column_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[column_letter].width = item['column_width']

            # 如果表头项有主标题，并且没有子标题或没有指定合并结束位置，则写入主标题到第一行
            if 'main' in item and ('sub' not in item or 'merge_start' in item):
                worksheet.cell(row=1, column=col_idx, value=item['main'])
            # 如果表头项有子标题，则写入子标题到第二行
            if 'sub' in item:
                worksheet.cell(row=2, column=col_idx, value=item['sub'])

            # 如果表头项只有主标题没有子标题，则合并第一行和第二行的对应列
            if 'main' in item and 'sub' not in item:
                worksheet.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            # 如果表头项有主标题、子标题，并且指定了合并结束位置，则合并主标题所在的单元格
            if 'main' in item and 'sub' in item and 'merge_start' in item:
                merge_end = len([i for i in self.headers if i['main'] == item['main']]) + col_idx - 1
                worksheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=merge_end)

            # 如果表头项指定了需要合并操作的列，则将该列索引添加到need_merged_column列表中
            if 'need_merged_column' in item and item['need_merged_column']:
                self.need_merged_column.append(col_idx)

    def _apply_styles(self, worksheet, df):
        """应用样式（自动处理所有列）"""
        total_rows = len(df) + 2  # 数据行数 + 标题行
        total_cols = len(self.headers)

        # 为所有单元格应用样式
        for row in range(1, total_rows + 1):
            for col in range(1, total_cols + 1):
                cell = worksheet.cell(row=row, column=col)

                # 设置单元格的数据类型为字符串
                cell.data_type = 's'
                cell.number_format = '@'

                # 确保单元格值为字符串
                if cell.value is not None and not isinstance(cell.value, str):
                    cell.value = str(cell.value)

                # 所有单元格应用边框
                cell.border = self.thin_border

                # 前两行（标题行）应用特殊样式
                if row <= 2:
                    cell.font = self.bold_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_alignment
                else:
                    # 数据行应用居中对齐
                    cell.alignment = self.center_alignment

    def process_merge(self, ws, data_cache, start_row, max_row, cols, control_cols=None, is_linked=False):
        """
        通用列合并函数（独立列或受联动控制的列）

        参数:
            ws (Worksheet): openpyxl 工作表对象
            data_cache (dict): {(row, col): value}，缓存的原始单元格数据，防止 merge 后变 None
            start_row (int): 数据起始行（通常跳过表头）
            max_row (int): 数据最大行
            cols (list[int]): 需要合并的列列表
            control_cols (list[int]): 联动列控制的列索引（仅 is_linked=True 时有效）
            is_linked (bool): 是否受控制列联动（True: 仅在控制列值相同时合并）
        """
        control_cols = control_cols or []

        for col in cols:
            merge_start = start_row
            in_group = False

            for row in range(start_row + 1, max_row + 1):
                current = data_cache[(row, col)]
                previous = data_cache[(row - 1, col)]

                # 判断是否属于同一组
                if is_linked:
                    control_same = all(
                        data_cache[(row, c)] == data_cache[(row - 1, c)]
                        for c in control_cols
                    )
                    same_group = control_same and current == previous and current not in [None, ""]
                else:
                    same_group = current == previous and current not in [None, ""]

                if same_group:
                    in_group = True
                    continue

                # 断组处理：只有连续组才执行 merge
                if in_group:
                    if row - merge_start >= 1:
                        ws.merge_cells(
                            start_row=merge_start, start_column=col,
                            end_row=row - 1, end_column=col
                        )
                        cell = ws.cell(row=merge_start, column=col)
                        cell.alignment = self.center_alignment
                        cell.border = self.thin_border

                merge_start = row
                in_group = False

            # 处理最后一组连续值
            if in_group and max_row - merge_start >= 1:
                ws.merge_cells(
                    start_row=merge_start, start_column=col,
                    end_row=max_row, end_column=col
                )
                cell = ws.cell(row=merge_start, column=col)
                cell.alignment = self.center_alignment
                cell.border = self.thin_border

    def merge_cells_full(self, ws, start_row=3, independent_cols=None, linked_cols=None, control_cols=None):
        """
        对 Excel 工作表进行多列合并

        参数:
            ws (Worksheet): openpyxl 工作表对象
            start_row (int): 数据起始行
            independent_cols (list[int]): 独立列（不受其他列控制）合并
            linked_cols (list[int]): 联动列（受 control_cols 控制）合并
            control_cols (list[int]): 控制列索引，只有当这些列值相同时才允许 linked_cols 合并
        """
        independent_cols = independent_cols or []
        linked_cols = linked_cols or []
        control_cols = control_cols or []

        max_row = ws.max_row

        # 缓存数据，避免合并后后破坏判断
        data_cache = {}
        for r in range(start_row, max_row + 1):
            for c in range(1, ws.max_column + 1):
                data_cache[(r, c)] = ws.cell(row=r, column=c).value

        # 执行独立列合并
        self.process_merge(ws, data_cache, start_row, max_row, independent_cols, is_linked=False)
        # 执行联动列合并
        self.process_merge(ws, data_cache, start_row, max_row, linked_cols, control_cols=control_cols, is_linked=True)
